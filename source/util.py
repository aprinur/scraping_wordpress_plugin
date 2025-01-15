from source.db_config import get_existing_table_class, create_db_table
from source import Session, engine, DATE, sql_reserved_keyword
from openpyxl.styles import Alignment, Font
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import inspect, text
from openpyxl import load_workbook
import openpyxl
import keyword
import pandas
import os


def save_as_file_input(confirm: bool = True, tablename: str = None) -> str | None:
    """ Function to get argument for export file from database table

    Args:
        confirm (bool): confirmation from the requesting function to save as file or not
        tablename (str): name of table from database to export as xlsx and csv

    Returns:
        str | None
    """

    while confirm:
        display_table()
        table_name = input('Enter table name: ')
        if not table_name.strip():
            print('Table name cannot be empty')
            continue
        if not check_table_exists(table_name):
            print("Table name doesn't exist in the database, re enter the correct table name")
            continue
        filename = input('Enter file name (optional): ') or 'WordPress plugins'
        sheet_title = input('Enter title for excel sheet (optional): ') or 'WordPress Plugin Scraping Result'
        sheet_desc = input(
            'Enter desc for excel sheet (optional): ') or f'This file exported from table {table_name} in the database'
        return filename, sheet_title, sheet_desc, table_name

    if not confirm:
        filename = input('Enter file name (optional): ') or 'WordPress plugins'
        sheet_title = input('Enter title for excel sheet (optional): ') or 'WordPress Plugin Scraping Result'
        sheet_desc = input('Enter desc for excel sheet (optional): ') or f'This file exported from table {tablename} in the database'
        return filename, sheet_title, sheet_desc, None


def input_pages() -> tuple:
    """ Getting certain parameter to scrape

        Args: None

        Returns:
            tuple: A tuple containing:
                - int: The number of pages to scrape.
                - str: The name of the table.
                - Type[WP_PLugins]: The class representing the existing table structure.
    """
    while True:
        pages = input('Enter total page to scrape (empty = all): ') or '2949'

        if pages.isdigit():
            pages = int(pages)
        else:
            print('Page must be a number')
            continue
        break

    while True:
        table_name = input('Enter table name: ').lower()
        valid = table_name_validator(table_name)

        if not valid:
            print('Re-enter table name')
            continue

        if check_table_exists(table_name):
            while True:
                add = input('Table already exist, do you want to append data to the existing table (y/n): ').lower()
                if add not in ['y', 'n']:
                    print("Choose only y or n")
                    continue

                if add == 'y':
                    try:
                        table_class = get_existing_table_class(table_name)
                        print(f'Appending data to the existing table: {table_name}.')
                        return pages, table_name, table_class
                    except ValueError as e:
                        print(f'Error on get existing table: {e}')
                        traceback.print_exc()
                        return
                else:
                    print('Enter the new table name!')
                    break
        else:
            table_class = create_db_table(table_name)
            print(f'Table {table_name} has created ')
            return pages, table_name, table_class


def insert_to_db(result, table_class) -> None:
    """Insert a new record into the database table.

        Args:
            result (dict): A dictionary containing the data to be inserted into the table.
            table_class (Type[WP_PLugins]): The class representing the table structure where the data will be inserted.

        Returns:
            None: This function does not return a value. It prints a success message or an error message if an exception occurs.
        """
    try:
        with Session() as session:
            table_instance = table_class(**result)
            session.add(table_instance)
            session.commit()
            print('Insert to db success')

    except Exception as e:
        print(f'Error {e}')


def check_db(data, table_class) -> bool:
    """Check whether a specific value exists in a certain table.

        Args:
            data (dict): A dictionary containing the value to check, expected to have a "Name" key.
            table_class (Type[WP_PLugins]): The class representing the table structure to query.

        Returns:
            bool: True if the value exists in the table, False otherwise.
    """
    with Session() as session:
        exist = session.query(table_class).filter_by(Name=data["Name"]).first()
        return exist is not None


def save_to_file(filename: str, sheet_title: str, sheet_desc: str, table_name: str = None) -> None:
    """ Function to export table from database into xlsx and csv file

        Args:
            filename (str): name of exported file
            sheet_title (str): title of the sheet from the file
            sheet_desc (str): description of the sheet, placed under title
            table_name (str): name of table from the database to export

        Returns: None
    """
    query = f'SELECT * FROM {table_name}'
    df = pandas.read_sql_query(query, engine)

    save_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(save_dir, exist_ok=True)

    excel_path = os.path.join(save_dir, f'{filename}_{DATE}.xlsx')
    csv_path = os.path.join(save_dir, f'{filename}_{DATE}.csv')

    with pandas.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, sheet_name='Sheet_1')
        worksheet = writer.sheets['Sheet_1']

        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet['A1'] = sheet_title
        worksheet['A2'] = sheet_desc

        worksheet['A1'].font = Font(name='Times New Roman', size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A2'].alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=4, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_col=5, max_col=5, min_row=4, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_col=6, max_col=6, min_row=4, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    df.to_csv(csv_path, index=False)
    print(f"\nSaved to file as: {filename}_{DATE}")
    return


def table_name_validator(tablename: str) -> bool:
    """ Function to validate value of tablename whether it is empty string,
    not a valid identifier or a reserved keyword list in python or sql

        Args:
            tablename (str): name of table to validate

        Returns: bool
    """
    if not tablename.strip():
        print('Table name cannot be empty')
        return False

    if not tablename.isidentifier():
        print('Table name must be a valid identifier')
        return False

    if tablename in keyword.kwlist or tablename.lower() in sql_reserved_keyword:
        print(f"Table name '{tablename}' is a reserved keyword and cannot be used")
        return False

    return True


def check_table_exists(tablename: str, engine_=engine) -> bool:
    """ Function to check exsistance of a table inside database

        Args:
            tablename (str): tablename that might be inside database
            engine_: an instance of SQLAlchemy's Engine

        Returns: bool
    """

    inspector = inspect(engine_)
    return tablename in inspector.get_table_names()


def user_input_del_table() -> str | bool:
    """ Function to get table name to delete

        arg: None

        Returns: str | bool
    """
    display_table()

    while True:
        table = input('\nEnter table name (empty = cancel): ')
        if not table.strip():
            print('Action aborted')
            return False
        if not check_table_exists(table):
            print('Table not found')
            continue

        return table.strip()


def delete_table_by_name(table_name: str) -> None:
    """
    Delete a table from WordPress_plugins.db database

        Args:
            table_name (str): name of table inside database to delete

        Returns: None
    """
    try:
        while table_name:
            confirm_deletion = input('Are you sure want to delete the table y/n: ').lower()
            if confirm_deletion not in ['y', 'n']:
                print('Enter y or n ')
                continue
            if confirm_deletion == 'n':
                print('Operation canceled')
                return

            with engine.connect() as connection:
                query = text('DROP TABLE IF EXISTS :table_name')
                connection.execute(text(query))
                print(f'Table {table_name} deletion report\t: Sucess')
            return
        return

    except SQLAlchemyError as e:
        print(f'''
        Table {table_name} deletion report\t: Failed
        Error when deleting table\t: {e}''')
        return


def display_table() -> None:
    """ Display all table from database WordPress_plugins.db

        Args: None

        Returns: None
    """
    try:
        with engine.connect():
            inspector = inspect(engine)
            tables = inspector.get_table_names()
            print('\nTables in WordPress_plugins.db')
            for num, table in enumerate(tables):
                print(f'{num + 1}. {table}')
    except SQLAlchemyError as e:
        print(f'Error while getting tables: {e}')
